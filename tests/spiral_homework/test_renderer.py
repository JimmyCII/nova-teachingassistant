# tests/spiral_homework/test_renderer.py
from openpyxl import load_workbook
from agent.tools.spiral_homework.models import Box, WeekSpec
from agent.tools.spiral_homework.renderer import render_xlsx

def _week():
    return WeekSpec(due_date="9/12", track="regular", boxes=[
        Box(day="Monday", type="computation", role="current", text="1 5/6 × 4 1/2"),
        Box(day="Monday", type="computation", role="review", text="246.75 ÷ 2.5 ="),
        Box(day="Tuesday", type="word_problem", role="review",
            text="A turtle swims 7.5 km in 3 hours. Unit rate?"),
        Box(day="Wednesday", type="brain_break", role="current", text=""),
        Box(day="Thursday", type="figure_placeholder", role="current",
            text="Plot the points.", figure_note="coordinate plane"),
    ])

def test_render_writes_xlsx_with_header_and_grid(tmp_path):
    out = render_xlsx(_week(), tmp_path / "wk.xlsx")
    assert out.exists()
    ws = load_workbook(out).active
    assert "Spiral Review Homework" in str(ws["B1"].value)
    assert "9/12" in str(ws["D1"].value)
    assert [ws.cell(row=2, column=c).value for c in range(1, 5)] == \
        ["Monday", "Tuesday", "Wednesday", "Thursday"]
    # Monday column (col 1) has the two computation problems stacked
    col1 = [ws.cell(row=r, column=1).value for r in range(3, 6)]
    assert any("1 5/6" in str(v) for v in col1)
    assert any("246.75" in str(v) for v in col1)
    # brain break label rendered
    allvals = [c.value for row in ws.iter_rows() for c in row if c.value]
    assert any("Brain Break" in str(v) for v in allvals)
    # figure placeholder note rendered
    assert any("coordinate plane" in str(v) for v in allvals)

def test_render_marks_borders(tmp_path):
    out = render_xlsx(_week(), tmp_path / "wk.xlsx")
    ws = load_workbook(out).active
    assert ws.cell(row=3, column=1).border.left.style == "thin"
