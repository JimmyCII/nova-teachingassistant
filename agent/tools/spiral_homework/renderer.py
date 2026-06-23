# agent/tools/spiral_homework/renderer.py
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from .models import Box, WeekSpec

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday"]

def _box_text(box: Box) -> str:
    if box.type == "brain_break":
        return "Brain Break!"
    text = box.text or ""
    if box.type == "figure_placeholder" or box.figure_note:
        text = f"{text}\n[figure: {box.figure_note or 'see teacher'}]".strip()
    return text

def render_xlsx(week: WeekSpec, out_path: str | Path) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Spiral Review"
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws["A1"] = "Name ____________"
    ws["B1"] = "Spiral Review Homework"; ws["B1"].alignment = center
    ws["D1"] = f"Due {week.due_date}"; ws["D1"].alignment = Alignment(horizontal="right")

    for i, day in enumerate(DAYS):
        c = ws.cell(row=2, column=i + 1, value=day)
        c.font = bold; c.alignment = center

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    by_day = {d: [b for b in week.boxes if b.day == d] for d in DAYS}
    max_rows = max((len(v) for v in by_day.values()), default=0)

    for col_i, day in enumerate(DAYS):
        for row_i, box in enumerate(by_day[day]):
            cell = ws.cell(row=3 + row_i, column=col_i + 1, value=_box_text(box))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col_i in range(1, 5):
        ws.column_dimensions[get_column_letter(col_i)].width = 26
    for r in range(3, 3 + max_rows):
        ws.row_dimensions[r].height = 90

    # Print as a single landscape page, like Karrie's originals.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    wb.save(out_path)
    return out_path
